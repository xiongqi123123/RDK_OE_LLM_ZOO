from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from typing import Any, Dict, List, Union

import numpy as np
import openpyxl
import openpyxl.utils
from openpyxl.styles import Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from .types import Colors, CompareResult, TensorInfo
from .utils import (
    calculate_cosine,
    debug_print,
    print_colored,
    to_json_serializable,
)


class BaseReporter:
    """Base class for generating comparison reports."""

    def __init__(self, mode_model_name: str, comparison_mode: str = "bc"):
        self.mode_model_name = mode_model_name
        self.comparison_mode = comparison_mode  # "bc" or "hbm"
        self.layers_detailed_data: defaultdict[str, Dict[str, List[CompareResult]]] = (
            defaultdict(dict)
        )
        self.outputs_detailed_data: defaultdict[str, Dict[str, List[CompareResult]]] = (
            defaultdict(dict)
        )

    @property
    def has_bc_mode(self) -> bool:
        """Check if BC mode is used (determines if layer data is processed)."""
        return self.comparison_mode == "bc"

    @property
    def all_modes(self) -> set[str]:
        """Get all modes from both layers and outputs data."""
        return set(self.layers_detailed_data.keys()).union(
            self.outputs_detailed_data.keys()
        )

    def _copy_data_to_reporter(self, reporter: "BaseReporter") -> None:
        """Copy data to another reporter instance."""
        reporter.layers_detailed_data = self.layers_detailed_data
        reporter.outputs_detailed_data = self.outputs_detailed_data
        reporter.comparison_mode = self.comparison_mode

    def generate_reports(self, output_dir: str = "./comparison_reports") -> None:
        """Generate all reports."""
        os.makedirs(output_dir, exist_ok=True)
        summary_stats = self._calculate_summary_statistics()

        console_reporter = ConsoleReporter(self.mode_model_name, self.comparison_mode)
        self._copy_data_to_reporter(console_reporter)
        console_reporter._generate_console_report(summary_stats)

        json_reporter = JsonReporter(self.mode_model_name, self.comparison_mode)
        self._copy_data_to_reporter(json_reporter)
        json_reporter._generate_json_report(output_dir, summary_stats)

        excel_reporter = ExcelReporter(self.mode_model_name, self.comparison_mode)
        self._copy_data_to_reporter(excel_reporter)
        excel_reporter._generate_excel_report(output_dir, summary_stats)

        # Additionally dump per-step cosines across layers and outputs
        # to a single CSV for detailed analysis
        if os.environ.get("LLM_VERIFIER_DEBUG"):
            self._dump_per_step_cosines(output_dir)

        # Print saved file paths
        json_path = os.path.abspath(
            os.path.join(output_dir, f"{self.mode_model_name}_comparison_report.json")
        )
        xlsx_path = os.path.abspath(
            os.path.join(output_dir, f"{self.mode_model_name}_comparison_report.xlsx")
        )
        print_colored("\nReports saved to:", Colors.WHITE)
        print_colored(f"  JSON: {json_path}", Colors.GREEN)
        print_colored(f"  Excel: {xlsx_path}", Colors.GREEN)

    def _calculate_summary_statistics(self) -> Dict[str, Any]:
        """Calculate summary statistics from all detailed data."""
        summary_stats: Dict[str, Any] = {}

        for mode in self.all_modes:
            mode_stats: Dict[str, Any] = {}

            # Only process layer results if BC mode was used
            if self.has_bc_mode:
                all_layer_results: List[CompareResult] = []
                for prompt_results in self.layers_detailed_data[mode].values():
                    all_layer_results.extend(prompt_results)

                # Only add layer_outputs stats if there are layer results
                if all_layer_results:
                    layer_cosines = [
                        r.cosine for r in all_layer_results if r.cosine is not None
                    ]
                    mode_stats["layer_outputs"] = {
                        "count": len(all_layer_results),
                        "valid_count": len(layer_cosines),
                        "mean_cosine": np.mean(layer_cosines)
                        if layer_cosines
                        else None,
                        "min_cosine": np.min(layer_cosines) if layer_cosines else None,
                        "max_cosine": np.max(layer_cosines) if layer_cosines else None,
                    }

            all_output_results: List[CompareResult] = []
            for prompt_results in self.outputs_detailed_data[mode].values():
                all_output_results.extend(prompt_results)

            # Only add final_outputs stats if there are output results
            if all_output_results:
                output_cosines = [
                    r.cosine for r in all_output_results if r.cosine is not None
                ]
                mode_stats["final_outputs"] = {
                    "count": len(all_output_results),
                    "valid_count": len(output_cosines),
                    "mean_cosine": np.mean(output_cosines) if output_cosines else None,
                    "min_cosine": np.min(output_cosines) if output_cosines else None,
                    "max_cosine": np.max(output_cosines) if output_cosines else None,
                }

            summary_stats[mode] = mode_stats

        return summary_stats

    @staticmethod
    def _infer_time_axis(shape: tuple[int, ...]) -> int | None:
        """Infer the time/chunk axis heuristically.

        Priority:
        1) If 3D and axis 1 equals 256 (common [B, T, H]) -> 1
        2) If penultimate axis equals 256 -> -2
        3) First axis with dim == 256
        Otherwise, None
        """
        if not shape:
            return None
        # Typical [B, T, H]
        if len(shape) >= 3 and shape[1] == 256:
            return 1
        # Many ops expose [*, T, *] with T at -2
        if len(shape) >= 2 and shape[-2] == 256:
            return -2
        for idx, dim in enumerate(shape):
            if dim == 256:
                return idx
        return None

    def _dump_per_step_cosines(self, output_dir: str) -> None:
        """Dump per-step (0..255) cosine similarity for all outputs/layers.

        Writes a single CSV file named
        "{model_name}_per_step_cosines.csv" with columns:
        mode,prompt_id,kind,name,step,cosine,torch_shape,bc_shape

        Only entries whose tensors expose a 256-sized time axis are included.
        """
        out_path = os.path.join(
            output_dir,
            f"{self.mode_model_name}_per_step_cosines.csv",
        )
        try:
            with open(out_path, "w", encoding="utf-8") as f:
                header = "mode,prompt_id,kind,name,step,cosine,torch_shape,bc_shape\n"
                f.write(header)

                # Dump outputs first
                for mode, prompts in self.outputs_detailed_data.items():
                    for prompt_id, results in prompts.items():
                        for comp in results:
                            t_info = comp.torch_info
                            b_info = comp.bc_info
                            if not t_info or not b_info:
                                continue
                            if t_info.data is None or b_info.data is None:
                                continue
                            axis = self._infer_time_axis(t_info.shape)
                            if axis is None:
                                continue
                            time_len = t_info.shape[axis]
                            for step in range(time_len):
                                t_slice = np.take(t_info.data, step, axis=axis)
                                b_slice = np.take(b_info.data, step, axis=axis)
                                cos = calculate_cosine(t_slice, b_slice)
                                cos_str = "" if cos is None else f"{cos:.6f}"
                                row = (
                                    f"{mode},{prompt_id},output,"
                                    f"{comp.name},{step},{cos_str},"
                                    f'"{t_info.shape}","{b_info.shape}"\n'
                                )
                                f.write(row)

                # Dump layers (BC mode only typically has layer data)
                for mode, prompts in self.layers_detailed_data.items():
                    for prompt_id, results in prompts.items():
                        for comp in results:
                            t_info = comp.torch_info
                            b_info = comp.bc_info
                            if not t_info or not b_info:
                                continue
                            if t_info.data is None or b_info.data is None:
                                continue
                            axis = self._infer_time_axis(t_info.shape)
                            if axis is None:
                                continue
                            time_len = t_info.shape[axis]
                            for step in range(time_len):
                                t_slice = np.take(t_info.data, step, axis=axis)
                                b_slice = np.take(b_info.data, step, axis=axis)
                                cos = calculate_cosine(t_slice, b_slice)
                                cos_str = "" if cos is None else f"{cos:.6f}"
                                row = (
                                    f"{mode},{prompt_id},layer,"
                                    f"{comp.name},{step},{cos_str},"
                                    f'"{t_info.shape}","{b_info.shape}"\n'
                                )
                                f.write(row)
        except Exception:
            # Best-effort debug dump; do not break reporting
            pass


class ConsoleReporter(BaseReporter):
    """Reporter for console output."""

    def __init__(self, mode_model_name: str, comparison_mode: str = "bc"):
        super().__init__(mode_model_name, comparison_mode)

    def _generate_console_report(self, summary_stats: Dict[str, Any]) -> None:
        """Print summary report."""
        print_colored("\nModel Comparison Report", Colors.BOLD)

        for mode in self.all_modes:
            stats = summary_stats.get(mode, {})
            print_colored(f"\nMode: {mode}", Colors.CYAN + Colors.BOLD)
            print_colored("-" * 60, Colors.CYAN)

            self._print_comparison_tables(mode)

            # Only print summary when BC mode was used and layer comparison data exists
            if self.has_bc_mode and self._aggregate_layers_data(mode):
                print_colored(
                    f"\n{'=' * 10} Summary Statistics for {mode} {'=' * 10}",
                    Colors.BLUE + Colors.BOLD,
                )

                layer_s = stats.get("layer_outputs", {})
                print_colored("Layer Outputs Summary:", Colors.BLUE)
                self._print_plain_stats(layer_s)

    def _print_plain_stats(self, stats: Dict[str, Any]) -> None:
        """Print statistics in plain text."""
        items = [
            ("Total", stats.get("count")),
            (
                "Valid",
                stats.get("valid_count"),
            ),
            (
                "Mean CosineSimilarity",
                stats.get("mean_cosine"),
            ),
            (
                "Min CosineSimilarity",
                stats.get("min_cosine"),
            ),
            (
                "Max CosineSimilarity",
                stats.get("max_cosine"),
            ),
        ]
        for k, v in items:
            v_str = f"{v:.6f}" if isinstance(v, (float, np.floating)) else str(v)
            print_colored(f"  {k}: {v_str}", Colors.WHITE)

    def _print_comparison_tables(self, mode: str) -> None:
        """Print layer and output comparison tables for a mode."""
        # Dynamic column names based on comparison mode
        comparison_mode_upper = self.comparison_mode.upper()
        if self.has_bc_mode:
            layer_data = self._aggregate_layers_data(mode)
            if layer_data:
                print_colored(
                    f"\n{'=' * 10} Each layer outputs comparison ({mode}) {'=' * 10}",
                    Colors.MAGENTA + Colors.BOLD,
                )
                layer_headers = [
                    "Layer",
                    "OriginalShape",
                    f"{comparison_mode_upper}Shape",
                    "OriginalDtype",
                    f"{comparison_mode_upper}Dtype",
                    "CosineSimilarity",
                ]
                self._print_table(layer_headers, layer_data)

        output_data = self._aggregate_data(mode, "outputs", "output")
        if output_data:
            print_colored(
                f"\n{'=' * 10} Model output tensor comparison ({mode}) {'=' * 10}",
                Colors.MAGENTA + Colors.BOLD,
            )
            output_headers = [
                "OutputName",
                "OriginalShape",
                f"{comparison_mode_upper}Shape",
                "OriginalDtype",
                f"{comparison_mode_upper}Dtype",
                "CosineSimilarity",
            ]
            self._print_table(output_headers, output_data)

    def _print_table(self, headers: List[str], data: List[Dict[str, Any]]) -> None:
        """Pretty-print a table with box-drawing borders."""
        if not data:
            print_colored("No data to display", Colors.YELLOW)
            return

        # Create custom field mapping for dynamic headers
        comparison_mode_upper = self.comparison_mode.upper()
        field_map: Dict[str, str] = {}
        for header in headers:
            if header == "Layer":
                field_map[header] = "layer"
            elif header == "OutputName":
                field_map[header] = "output"
            elif header == "OriginalShape":
                field_map[header] = "shape_torch"
            elif header == f"{comparison_mode_upper}Shape":
                field_map[header] = "shape_bc"
            elif header == "OriginalDtype":
                field_map[header] = "dtype_torch"
            elif header == f"{comparison_mode_upper}Dtype":
                field_map[header] = "dtype_bc"
            elif header == "CosineSimilarity":
                field_map[header] = "cosine"
            else:
                # Fallback to original behavior
                field_map[header] = header.lower().replace(" ", "_")

        col_widths: List[int] = []
        for header in headers:
            max_len = len(header)
            field = field_map.get(header, header)
            for row in data:
                value_str = str(row.get(field, "N/A"))
                max_len = max(max_len, len(value_str))
            col_widths.append(max_len + 2)

        def make_border(left: str, mid: str, right: str) -> str:
            return left + mid.join("─" * w for w in col_widths) + right

        print_colored(make_border("┌", "┬", "┐"), Colors.WHITE)

        header_row = (
            "│"
            + "│".join(f" {h:<{col_widths[i] - 1}}" for i, h in enumerate(headers))
            + "│"
        )
        print_colored(header_row, Colors.WHITE + Colors.BOLD)

        print_colored(make_border("├", "┼", "┤"), Colors.WHITE)

        for row in data:
            line = "│"
            for i, header in enumerate(headers):
                field = field_map.get(header, header)
                val = str(row.get(field, "N/A"))
                line += f" {val:<{col_widths[i] - 1}}│"
            print_colored(line, Colors.WHITE)

        print_colored(make_border("└", "┴", "┘"), Colors.WHITE)

    def _aggregate_data(
        self, mode: str, data_source: str, key_name: str
    ) -> List[Dict[str, Any]]:
        """Generic method to aggregate comparison data across prompts for a mode.

        Args:
            mode: The comparison mode
            data_source: Either 'layers' or 'outputs'
            key_name: Either 'layer' or 'output'
        """
        aggregated: List[Dict[str, Any]] = []
        cosine_map: Dict[str, List[float]] = defaultdict(list)

        # Select the appropriate data source
        if data_source == "layers":
            source_data = self.layers_detailed_data.get(mode, {})
        else:  # outputs
            source_data = self.outputs_detailed_data.get(mode, {})

        for prompt_results in source_data.values():
            for comp in prompt_results:
                entry = comp.to_summary_dict()
                if data_source == "outputs":
                    entry["output"] = entry.pop("layer")
                name = entry[key_name]
                if not any(e[key_name] == name for e in aggregated):
                    aggregated.append(entry)
                if comp.cosine is not None:
                    cosine_map[name].append(comp.cosine)

        for entry in aggregated:
            name = entry[key_name]
            if cosine_map[name]:
                entry["cosine"] = f"{np.mean(cosine_map[name]):.6f}"

        return aggregated

    def _aggregate_layers_data(self, mode: str) -> List[Dict[str, Any]]:
        """Aggregate layer comparison data across prompts for a mode."""
        return self._aggregate_data(mode, "layers", "layer")


class JsonReporter(BaseReporter):
    """Reporter for JSON files."""

    def __init__(self, mode_model_name: str, comparison_mode: str = "bc"):
        super().__init__(mode_model_name, comparison_mode)

    def _generate_json_report(
        self, output_dir: str, summary_stats: Dict[str, Any]
    ) -> None:
        """Save JSON report."""
        report_data: Dict[str, Any] = {
            "model_name": self.mode_model_name,
            "summary": summary_stats,
            "detailed": {},
        }

        for mode in self.all_modes:
            detailed_mode_data: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
            report_data["detailed"][mode] = detailed_mode_data

            # Only process layer_outputs if BC mode was used
            if (
                self.has_bc_mode
                and mode in self.layers_detailed_data
                and self.layers_detailed_data[mode]
            ):
                layer_data = {}
                for prompt_id, results in self.layers_detailed_data[mode].items():
                    if results:  # Only add non-empty results
                        layer_data[prompt_id] = [
                            result.to_detailed_dict() for result in results
                        ]
                if layer_data:  # Only add layer_outputs if there's data
                    detailed_mode_data["layer_outputs"] = layer_data

            # Only add final_outputs if there's actual output data
            if mode in self.outputs_detailed_data and self.outputs_detailed_data[mode]:
                output_data = {}
                for prompt_id, results in self.outputs_detailed_data[mode].items():
                    if results:  # Only add non-empty results
                        output_data[prompt_id] = [
                            result.to_detailed_dict() for result in results
                        ]
                if output_data:  # Only add final_outputs if there's data
                    detailed_mode_data["final_outputs"] = output_data

        json_data = to_json_serializable(report_data)

        json_path = os.path.join(
            output_dir, f"{self.mode_model_name}_comparison_report.json"
        )
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)


class ExcelReporter(BaseReporter):
    """Reporter for Excel files."""

    def __init__(self, mode_model_name: str, comparison_mode: str = "bc"):
        super().__init__(mode_model_name, comparison_mode)

    def _generate_excel_report(
        self, output_dir: str, summary_stats: Dict[str, Any]
    ) -> None:
        """Save Excel report."""
        excel_path = os.path.join(
            output_dir, f"{self.mode_model_name}_comparison_report.xlsx"
        )
        workbook = openpyxl.Workbook()

        if workbook.active is not None:
            workbook.remove(workbook.active)

        summary_sheet = workbook.create_sheet("Summary", 0)
        self._write_summary_sheet(summary_sheet, summary_stats)

        sheet_index = 1
        for mode in self.all_modes:
            # Get all prompt/image IDs for this mode
            prompt_ids = set()
            if mode in self.outputs_detailed_data:
                prompt_ids.update(self.outputs_detailed_data[mode].keys())
            if self.has_bc_mode and mode in self.layers_detailed_data:
                prompt_ids.update(self.layers_detailed_data[mode].keys())

            for prompt_id in prompt_ids:
                sheet_name = f"{mode}_{prompt_id}"[:31]
                detail_sheet = workbook.create_sheet(sheet_name, sheet_index)

                # Get layer and output results
                layer_results = []
                if self.has_bc_mode and mode in self.layers_detailed_data:
                    layer_results = self.layers_detailed_data[mode].get(prompt_id, [])

                output_results = []
                if mode in self.outputs_detailed_data:
                    output_results = self.outputs_detailed_data[mode].get(prompt_id, [])

                self._write_detailed_sheet(
                    detail_sheet,
                    mode,
                    prompt_id,
                    layer_results,
                    output_results,
                )
                sheet_index += 1

        workbook.save(excel_path)
        workbook.close()

    def _write_summary_sheet(
        self, sheet: Worksheet, summary_stats: Dict[str, Any]
    ) -> None:
        """Write summary worksheet."""
        # Use dynamic column count (F = 6 columns is reasonable for summary)
        summary_cols = 6
        last_col = openpyxl.utils.get_column_letter(summary_cols)

        sheet["A1"] = "Model Comparison Report"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet.merge_cells(f"A1:{last_col}1")

        row = 3
        for mode, mode_stats in summary_stats.items():
            sheet[f"A{row}"] = f"Mode: {mode}"
            sheet[f"A{row}"].font = Font(size=14, bold=True)
            sheet.merge_cells(f"A{row}:{last_col}{row}")
            row += 1

            # Only write layer outputs if they exist
            if "layer_outputs" in mode_stats:
                row = self._write_stats_section(
                    sheet, row, "Layer Outputs", mode_stats["layer_outputs"]
                )
                row += 1

            # Only write final outputs if they exist
            if "final_outputs" in mode_stats:
                row = self._write_stats_section(
                    sheet, row, "Final Outputs", mode_stats["final_outputs"]
                )
                row += 1

            row += 3

        for i in range(1, sheet.max_column + 1):
            column_letter = openpyxl.utils.get_column_letter(i)
            max_length = 0
            for cell in sheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 4, 60)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _write_stats_section(
        self, sheet: Worksheet, row: int, section_name: str, stats: Dict[str, Any]
    ) -> int:
        """Write a section of statistics to the worksheet."""
        # Use the same column count as summary sheet
        summary_cols = 6
        last_col = openpyxl.utils.get_column_letter(summary_cols)

        sheet[f"A{row}"] = section_name
        sheet[f"A{row}"].font = Font(size=12, bold=True)
        sheet.merge_cells(f"A{row}:{last_col}{row}")
        row += 1

        items = [
            ("Total", stats.get("count")),
            ("Valid", stats.get("valid_count")),
            ("Mean cosine", stats.get("mean_cosine")),
            ("Min", stats.get("min_cosine")),
            ("Max", stats.get("max_cosine")),
        ]
        for k, v in items:
            v_str = f"{v:.6f}" if isinstance(v, (float, np.floating)) else str(v)
            sheet[f"A{row}"].value = k
            sheet[f"B{row}"].value = v_str
            sheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 15
            sheet.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 15
            row += 1
        return row

    def _write_detailed_sheet(
        self,
        sheet: Worksheet,
        mode: str,
        prompt_id: str,
        layer_results: List[CompareResult],
        output_results: List[CompareResult],
    ) -> None:
        """Write a detailed comparison sheet for a specific prompt."""
        # Define headers with dynamic comparison mode
        comparison_mode_upper = self.comparison_mode.upper()
        headers = [
            "OutputName",
            "OriginalShape",
            f"{comparison_mode_upper}Shape",
            "OriginalDtype",
            f"{comparison_mode_upper}Dtype",
            "CosineSimilarity",
        ]

        # Dynamic column range based on header count
        last_col = openpyxl.utils.get_column_letter(len(headers))

        sheet["A1"] = f"Model Comparison - {mode} - {prompt_id}"
        sheet["A1"].font = Font(size=14, bold=True)
        sheet.merge_cells(f"A1:{last_col}1")

        # Write header that correctly corresponds to the data structure
        sheet.append(headers)

        # Write data
        for result in layer_results:
            sheet.append(list(result.to_detailed_dict().values()))

        for result in output_results:
            sheet.append(list(result.to_detailed_dict().values()))

        sheet.freeze_panes = "A2"

        # Add borders to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )


class ComparisonReporter(BaseReporter):
    """
    Compare the outputs of the torch and bc models and generate reports.
    This class now delegates report generation to specialized classes.
    """

    def __init__(self, mode_model_name: str, comparison_mode: str = "bc") -> None:
        """Initialize the comparison reporter."""
        super().__init__(mode_model_name, comparison_mode)

    def compare_inference_results(
        self,
        results: Dict[str, List[Union[Dict[str, TensorInfo], List[TensorInfo], None]]],
        compare_mode: str,
        prompt_id: str | None = None,
        image_id: str | None = None,
        last_valid_step: int | None = None,
    ) -> None:
        """
        Compare inference results in the new dictionary format.

        Args:
            results: Dictionary with "llm" and "vlm" keys containing inference results.
            compare_mode: Comparison mode ("bc" or "hbm").
            prompt_id: Identifier for the text prompt, required for LLM results.
            image_id: Identifier for the image input, required for VLM results.
        """
        # Track the compare mode being used
        self.comparison_mode = compare_mode

        # Handle LLM results
        llm_results = results.get("llm", [])
        if llm_results:
            if not prompt_id:
                raise ValueError("A prompt_id must be provided for LLM results.")
            self._handle_model_results(
                llm_results, compare_mode, "llm", prompt_id, last_valid_step
            )

        # Handle VLM results
        vlm_results = results.get("vlm", [])
        if vlm_results:
            if not image_id:
                raise ValueError("An image_id must be provided for VLM results.")
            self._handle_model_results(vlm_results, compare_mode, "vlm", image_id, None)
        elif "vlm" in results:
            print("No VLM results found, skipping VLM comparison")

    def _handle_model_results(
        self,
        model_results: List[Union[Dict[str, TensorInfo], List[TensorInfo], None]],
        compare_mode: str,
        model_type: str,  # "llm" or "vlm"
        result_id: str,  # prompt_id for LLM, image_id for VLM
        last_valid_step: int | None = None,
    ) -> None:
        """Generic handler for model comparison results."""
        if compare_mode == "bc":
            if len(model_results) >= 4:
                torch_out, bc_out, torch_layers, bc_layers = model_results[:4]
                if torch_out is not None and bc_out is not None:
                    print(
                        (
                            f"Comparing {model_type.upper()} outputs for "
                            f"{result_id} in {compare_mode.upper()} mode..."
                        )
                    )
                    self.compare_layers_input(
                        model_type,
                        result_id,
                        torch_out,
                        bc_out,
                        torch_layers,
                        bc_layers,
                        last_valid_step=last_valid_step,
                    )
                else:
                    debug_print(
                        (
                            f"Warning: Some {model_type.upper()} "
                            f"{compare_mode.upper()} outputs for {result_id} "
                            "are None, skipping comparison"
                        )
                    )
            elif len(model_results) >= 2:
                debug_print(
                    (
                        f"Warning: {compare_mode.upper()} backend not available "
                        f"for {model_type.upper()}, only Torch "
                        f"{model_type.upper()} results collected"
                    )
                )
            else:
                debug_print(
                    (
                        f"Warning: Insufficient {model_type.upper()} results for "
                        f"{compare_mode.upper()} comparison"
                    )
                )
        elif compare_mode == "hbm":
            if len(model_results) >= 2:
                torch_out, hbm_out = model_results[:2]
                if torch_out is not None and hbm_out is not None:
                    print(
                        (
                            f"Comparing {model_type.upper()} outputs for "
                            f"{result_id} in {compare_mode.upper()} mode..."
                        )
                    )
                    self.compare_outputs_only(
                        model_type,
                        result_id,
                        torch_out,
                        hbm_out,
                        "hbm",
                        last_valid_step=last_valid_step,
                    )
                else:
                    debug_print(
                        (
                            f"Warning: Some {model_type.upper()} "
                            f"{compare_mode.upper()} outputs for {result_id} "
                            "are None, skipping comparison"
                        )
                    )
            else:
                debug_print(
                    (
                        f"Warning: Insufficient {model_type.upper()} results for "
                        f"{compare_mode.upper()} comparison"
                    )
                )

    def _get_last_valid_step_index(self, torch_info: TensorInfo) -> int | None:
        """Return the last step index within chunk_size-1 range for prefill.

        New rule: always take the final prefill position within its chunk.
        If the time axis length is L, return L - 1.
        """
        if torch_info is None or torch_info.data is None or not torch_info.shape:
            return None

        axis = self._infer_time_axis(torch_info.shape)
        if axis is None:
            return None

        try:
            time_len = int(torch_info.shape[axis])
            if time_len <= 0:
                return None
            return time_len - 1
        except Exception:
            return None

    def _slice_to_last_valid_step(
        self,
        torch_info: TensorInfo,
        bc_info: TensorInfo,
        last_valid_step: int | None = None,
    ) -> tuple[np.ndarray | None, np.ndarray | None]:
        """Return (torch_slice, bc_slice) at the last valid prefill step.

        - Determine index from torch side via `_get_last_valid_step_index`.
        - Slice both tensors along their inferred time axes.
        - If slicing is not possible, return original arrays.
        """
        if torch_info is None or bc_info is None:
            return (
                torch_info.data if torch_info else None,
                bc_info.data if bc_info else None,
            )

        if torch_info.data is None or bc_info.data is None:
            return torch_info.data, bc_info.data

        t_axis = self._infer_time_axis(torch_info.shape)
        b_axis = self._infer_time_axis(bc_info.shape)
        if t_axis is None or b_axis is None:
            return torch_info.data, bc_info.data

        idx = (
            int(last_valid_step)
            if last_valid_step is not None
            else self._get_last_valid_step_index(torch_info)
        )
        if idx is None:
            return torch_info.data, bc_info.data

        # Clamp index to each side's time length
        t_len = torch_info.shape[t_axis]
        b_len = bc_info.shape[b_axis]
        if t_len <= 0 or b_len <= 0:
            return torch_info.data, bc_info.data

        safe_idx = max(0, min(idx, t_len - 1, b_len - 1))

        try:
            t_slice = np.take(torch_info.data, safe_idx, axis=t_axis)
            b_slice = np.take(bc_info.data, safe_idx, axis=b_axis)
            return t_slice, b_slice
        except Exception:
            return torch_info.data, bc_info.data

    def compare_layers_input(
        self,
        mode: str,
        prompt_id: str,
        torch_outputs: Union[Dict[str, TensorInfo], List[TensorInfo]],
        bc_outputs: Union[Dict[str, TensorInfo], List[TensorInfo]],
        torch_layer_outputs: Union[Dict[str, TensorInfo], List[TensorInfo]],
        bc_layer_outputs: Union[Dict[str, TensorInfo], List[TensorInfo]],
        last_valid_step: int | None = None,
    ) -> None:
        """Compare the layers and outputs of the torch and bc models."""
        # This method is used for BC mode comparisons
        self.comparison_mode = "bc"

        output_comparisons = self._compare_outputs(
            torch_outputs, bc_outputs, last_valid_step
        )
        layer_comparisons = self._compare_tensors(
            torch_layer_outputs, bc_layer_outputs, last_valid_step
        )

        self.outputs_detailed_data[mode][prompt_id] = output_comparisons
        self.layers_detailed_data[mode][prompt_id] = layer_comparisons

    def compare_outputs_only(
        self,
        mode: str,
        prompt_id: str,
        torch_outputs: Union[Dict[str, TensorInfo], List[TensorInfo]],
        comparison_outputs: Union[Dict[str, TensorInfo], List[TensorInfo]],
        comparison_type: str = "bc",
        last_valid_step: int | None = None,
    ) -> None:
        """Compare only the final outputs.

        This method aligns outputs purely by return order and only compares the
        first output tensor (index 0) from both sides. For HBM comparison we are
        only interested in ``output_0``; for BC comparison this keeps the same
        lightweight behaviour.
        """
        # Track the compare mode being used
        self.comparison_mode = comparison_type

        # Sequential comparison (first tensor only) — works for both BC and HBM.
        output_results = self._compare_outputs(
            torch_outputs, comparison_outputs, last_valid_step
        )

        self.outputs_detailed_data[mode][prompt_id] = output_results

    def _find_matching_bc_key(
        self,
        torch_layer_name: str,
        torch_info: TensorInfo,
        bc_layers_outputs: Dict[str, TensorInfo],
        last_valid_step: int | None = None,
    ) -> str | None:
        """Locate the best-matched BC tensor key for the given torch layer.

        The strategy mirrors the previous standalone util but is now encapsulated
        within the reporter:
        1. Filter BC keys by layer name substring.
        2. Ensure the element count (shape product) matches.
        3. Choose the candidate with the highest cosine similarity at the
           selected time index when possible.
        """

        torch_pattern_name = torch_layer_name.replace("model.", "").replace(".", "_")
        torch_shape_mul = int(np.prod(torch_info.shape))

        candidates: list[str] = []
        for bc_key, bc_info in bc_layers_outputs.items():
            name_match = re.search(r'"([^\"]+)"[^\"]*$', bc_key)
            if not name_match:
                continue
            bc_op_name = name_match.group(1).replace(".", "_")
            if bc_op_name not in torch_pattern_name:
                continue
            if int(np.prod(bc_info.shape)) != torch_shape_mul:
                continue
            candidates.append(bc_key)

        best_key: str | None = None
        best_cosine: float = -1.0

        for key in candidates:
            bc_info = bc_layers_outputs[key]
            cosine_val: float | None = None
            if torch_info.data is not None and bc_info.data is not None:
                t_slice, b_slice = self._slice_to_last_valid_step(
                    torch_info, bc_info, last_valid_step
                )
                cosine_val = calculate_cosine(t_slice, b_slice)
            if cosine_val is not None:
                if cosine_val > best_cosine:
                    best_cosine = cosine_val
                    best_key = key
            else:
                best_key = key  # fallback when cosine cannot be calculated
        return best_key

    def _compare_tensors(
        self,
        torch_tensors: Union[Dict[str, TensorInfo], List[TensorInfo]],
        bc_tensors: Union[Dict[str, TensorInfo], List[TensorInfo]],
        last_valid_step: int | None = None,
    ) -> List[CompareResult]:
        """Compare torch and bc tensors, calculate cosine similarity."""
        comparisons: List[CompareResult] = []

        torch_dict = self._normalize_to_dict(torch_tensors, "torch")
        bc_dict = self._normalize_to_dict(bc_tensors, "bc")

        for torch_name, torch_info in torch_dict.items():
            compare_result = CompareResult(name=torch_name, torch_info=torch_info)

            bc_key = self._find_matching_bc_key(
                torch_name, torch_info, bc_dict, last_valid_step
            )
            if bc_key and bc_key in bc_dict:
                bc_info = bc_dict[bc_key]
                compare_result.set_bc_info(bc_info)
                # Only compare the last valid prefill step along time axis
                t_slice, b_slice = self._slice_to_last_valid_step(
                    torch_info, bc_info, last_valid_step
                )
                cosine_similarity = calculate_cosine(t_slice, b_slice)
                compare_result.cosine = cosine_similarity

            comparisons.append(compare_result)

        return comparisons

    def _compare_outputs(
        self,
        torch_outputs: Union[Dict[str, TensorInfo], List[TensorInfo]],
        bc_outputs: Union[Dict[str, TensorInfo], List[TensorInfo]],
        last_valid_step: int | None = None,
    ) -> List[CompareResult]:
        """Compare torch and bc outputs sequentially without name matching."""
        comparisons: List[CompareResult] = []

        torch_list = self._get_output_data(torch_outputs)
        bc_list = self._get_output_data(bc_outputs)

        max_len = min(len(torch_list), len(bc_list))
        for i in range(max_len):
            output_name = f"output_{i}"

            if i < len(torch_list):
                torch_info = torch_list[i]
                compare_result = CompareResult(name=output_name, torch_info=torch_info)

                if i < len(bc_list):
                    bc_info = bc_list[i]
                    compare_result.set_bc_info(bc_info)
                    # Only compare the last valid prefill step along time axis
                    t_slice, b_slice = self._slice_to_last_valid_step(
                        torch_info, bc_info, last_valid_step
                    )
                    cosine_similarity = calculate_cosine(t_slice, b_slice)
                    compare_result.cosine = cosine_similarity
                else:
                    continue

                comparisons.append(compare_result)

        return comparisons

    def _normalize_to_dict(
        self, tensors: Union[Dict[str, TensorInfo], List[TensorInfo]], prefix: str
    ) -> Dict[str, TensorInfo]:
        """Normalize tensors to dictionary format."""
        if isinstance(tensors, dict):
            return tensors
        elif isinstance(tensors, list):
            return {f"{prefix}_output_{i}": tensor for i, tensor in enumerate(tensors)}
        else:
            raise ValueError(f"Unsupported tensor type: {type(tensors)}")

    def _get_output_data(
        self, tensors: Union[Dict[str, TensorInfo], List[TensorInfo]]
    ) -> List[TensorInfo]:
        """Normalize tensors to list format."""
        if isinstance(tensors, list):
            if os.environ.get("LLM_VERIFIER_DEBUG"):
                return tensors
            else:
                return tensors[:1]
        elif isinstance(tensors, dict):
            if os.environ.get("LLM_VERIFIER_DEBUG"):
                return list(tensors.values())
            else:
                return list(tensors.values())[:1]
        else:
            raise ValueError(f"Unsupported tensor type: {type(tensors)}")
